from __future__ import annotations

import io
import logging
from datetime import date
from typing import Any

from fastapi import APIRouter, Response, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from .service import RevisionService
    from .redis_store import RedisStore
except ImportError:
    from service import RevisionService
    from redis_store import RedisStore

router = APIRouter()
service = RevisionService(RedisStore())

logger = logging.getLogger(__name__)


def _to_iso_string(value: Any) -> str:
    """Convert date/datetime to ISO string, else return string or empty"""
    if isinstance(value, date):
        return value.isoformat()
    return str(value) if value else ""


def style_header_cell(cell):
    """Apply header styling"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def style_data_cell(cell):
    """Apply data cell styling"""
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def build_filials_excel(filials: list[dict[str, Any]]) -> bytes:
    """Generate Excel file with filials data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Филиалы"

    # Headers
    headers = [
        "ID", "Название филиала", "Первая ревизия", "Предыдущая ревизия",
        "Следующая ревизия", "Статус", "Недостача", "Сумма недостачи", "Даты ревизий"
    ]

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_header_cell(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Adjust specific column widths
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["H"].width = 30

    # Write data rows
    for row_idx, filial in enumerate(filials, start=2):
        revision_dates_list = filial.get("revision_dates", [])
        revision_dates_str = ", ".join(
            _to_iso_string(d) for d in revision_dates_list if d
        ) if isinstance(revision_dates_list, list) else ""

        # Сумма недостачи по всем ревизиям
        revision_shortages = filial.get("revision_shortages", {})
        total_shortage = sum(float(v) for v in revision_shortages.values()) if revision_shortages else 0

        row_data = [
            filial.get("id", ""),
            filial.get("name", ""),
            _to_iso_string(filial.get("first_revision_date")),
            _to_iso_string(filial.get("previous_revision_date")),
            _to_iso_string(filial.get("next_revision_date")),
            filial.get("next_revision_status", "planned"),
            filial.get("shortage", 0),
            total_shortage,
            revision_dates_str,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_data_cell(cell)

            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    ws.freeze_panes = "B2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_holidays_excel(holidays: list[dict[str, Any]]) -> bytes:
    """Generate Excel file with holidays data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Праздники"

    headers = ["ID", "Дата", "Название праздника"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_header_cell(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    ws.column_dimensions["C"].width = 30

    for row_idx, holiday in enumerate(holidays, start=2):
        row_data = [
            holiday.get("id", ""),
            holiday.get("date", ""),
            holiday.get("name", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_data_cell(cell)

            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.get("/export/filials")
async def export_filials():
    """Export all filials to Excel"""
    try:
        filials = service.list_filials()
        excel_bytes = build_filials_excel(filials)

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="filials.xlsx"'
            }
        )
    except Exception as e:
        logger.error(f"Error exporting filials: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export filials: {str(e)}")


@router.get("/export/holidays")
async def export_holidays():
    """Export all holidays to Excel"""
    try:
        holidays = service.holidays_store.get_all_holidays()
        excel_bytes = build_holidays_excel(holidays)

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="holidays.xlsx"'
            }
        )
    except Exception as e:
        logger.error(f"Error exporting holidays: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export holidays: {str(e)}")
