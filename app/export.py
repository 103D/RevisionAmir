from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("revision.export")

# ---------- styles ----------
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0D6B4B", end_color="0D6B4B", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
MONEY_FORMAT = '#,##0.00'
ALT_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

# ---------- column config ----------
COLUMNS = [
    {"header": "ID", "field": "id", "width": 8},
    {"header": "Год", "field": "year", "width": 8},
    {"header": "Месяц", "field": "month", "width": 12},
    {"header": "Объект", "field": "project", "width": 25},
    {"header": "Ревизия", "field": "revision_info", "width": 20},
    {"header": "Дата", "field": "event_date", "width": 14},
    {"header": "День недели", "field": "weekday", "width": 14},
    {"header": "Тип проверки", "field": "inspection_type", "width": 20},
    {"header": "Статус", "field": "status", "width": 16},
    {"header": "План", "field": "amount_planned", "width": 14},
    {"header": "Факт", "field": "amount_actual", "width": 14},
]

# ---------- status colors ----------
STATUS_COLORS = {
    "проведено": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    "запланировано": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    "не проведено": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    "выполнено": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
}


def _format_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _get_status_fill(status: str) -> PatternFill | None:
    s = (status or "").lower()
    for key, fill in STATUS_COLORS.items():
        if key in s:
            return fill
    return None


def generate_excel(
    items: list[dict[str, Any]],
    title: str = "Расписание ревизий",
) -> io.BytesIO:
    """Generate an Excel file from schedule items and return as BytesIO."""
    logger.info("Generating Excel export with %d rows", len(items))

    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    # ---------- title row ----------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Arial", size=14, bold=True, color="0D6B4B")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ---------- subtitle ----------
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    subtitle_cell = ws.cell(
        row=2, column=1,
        value=f"Экспорт: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Записей: {len(items)}"
    )
    subtitle_cell.font = Font(name="Arial", size=9, italic=True, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # ---------- headers (row 4) ----------
    header_row = 4
    for col_idx, col_def in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_def["header"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

    ws.row_dimensions[header_row].height = 24

    # ---------- data rows ----------
    for row_idx, item in enumerate(items, start=header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0

        for col_idx, col_def in enumerate(COLUMNS, start=1):
            field = col_def["field"]
            value = _format_value(item.get(field))

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

            # alt row shading
            if is_alt:
                cell.fill = ALT_ROW_FILL

            # status coloring
            if field == "status" and isinstance(value, str):
                status_fill = _get_status_fill(value)
                if status_fill:
                    cell.fill = status_fill

            # money formatting
            if field in ("amount_planned", "amount_actual") and isinstance(value, (int, float)):
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # center align for specific columns
            if field in ("id", "year", "event_date", "weekday"):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---------- totals row ----------
    if items:
        total_row = header_row + len(items) + 1
        ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(name="Arial", size=10, bold=True)

        # sum planned
        planned_col = next(i for i, c in enumerate(COLUMNS, 1) if c["field"] == "amount_planned")
        planned_cell = ws.cell(
            row=total_row,
            column=planned_col,
            value=sum(item.get("amount_planned") or 0 for item in items)
        )
        planned_cell.font = Font(name="Arial", size=10, bold=True)
        planned_cell.number_format = MONEY_FORMAT
        planned_cell.alignment = Alignment(horizontal="right")
        planned_cell.border = THIN_BORDER

        # sum actual
        actual_col = next(i for i, c in enumerate(COLUMNS, 1) if c["field"] == "amount_actual")
        actual_cell = ws.cell(
            row=total_row,
            column=actual_col,
            value=sum(item.get("amount_actual") or 0 for item in items)
        )
        actual_cell.font = Font(name="Arial", size=10, bold=True)
        actual_cell.number_format = MONEY_FORMAT
        actual_cell.alignment = Alignment(horizontal="right")
        actual_cell.border = THIN_BORDER

    # ---------- freeze panes ----------
    ws.freeze_panes = f"A{header_row + 1}"

    # ---------- auto filter ----------
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(items)}"

    # ---------- print setup ----------
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("Excel export generated successfully, %d bytes", output.getbuffer().nbytes)
    return output